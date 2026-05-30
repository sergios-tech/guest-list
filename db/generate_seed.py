"""
Generate db/02_seed.sql from the source xlsx.

Logic:
  - 1 invitation row per xlsx row.
  - Status maps to the enum.
  - Accommodation is extracted from Napomena via regex.
  - For Odbijeno rows, Napomena becomes decline_reason.
  - When Napomena looks like a comma-separated list of names
    (and status is not Odbijeno), each name becomes an attendee.
  - Anything not parsed cleanly stays in the `notes` column.
"""
import re
import sys
import unicodedata
import uuid
from pathlib import Path
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
SRC = HERE / "Spisak gostiju za svadbu.xlsx"
OUT = HERE / "02_seed.sql"


def norm(value):
    """NFC-normalise + strip. Required because cells may contain pre-composed
    `đ` (U+0111) or its NFD form `d`+U+0335 — both look identical in editors
    but only the NFC form matches our dict keys. Untouched, NFD rows fall
    through to the default ('NIJE_POZVAN') and silently break stats."""
    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value)).strip()


STATUS_MAP = {
    norm(k): v
    for k, v in {
        "Nije pozvan":         "NIJE_POZVAN",
        "Pozvan":              "POZVAN",
        "Odbijeno":            "ODBIJENO",
        "Potvrđen dolazak":    "POTVRDJEN_DOLAZAK",
    }.items()
}

# Accommodation patterns (case-insensitive). Order matters: longer first.
# We also pre-strip 'potreban smeštaj' prefix because it always co-occurs
# with an accommodation token and creates orphan words.
PREFIX_PATTERNS = [
    re.compile(r"potreb(an|no|na)\s+sme[sš]taj\s+u?\s*", re.I),
    re.compile(r"sme[sš]taj\s+", re.I),
]
ACCOM_PATTERNS = [
    (re.compile(r"siesta\s+apartman",              re.I), "SIESTA_APARTMENT"),
    (re.compile(r"(siesta\s+jednokrevetna|jednokrevetna\s+siesta)", re.I), "SIESTA_SINGLE"),
    (re.compile(r"(siesta\s+dvokrevetna|dvokrevetna\s+siesta)",     re.I), "SIESTA_DOUBLE"),
    (re.compile(r"\bsiesti\b",                     re.I), "SIESTA_DOUBLE"),
    (re.compile(r"\baria\b",                       re.I), "ARIA"),
]

# Words that aren't real names — note categories
NON_NAME_WORDS = {
    "drugarice", "koleginice", "komšiluk", "komsiluk", "kuma",
    "jovana planira troje", "jovana planira samo odrasle",
    "potreban smeštaj u siesti", "potreban smestaj u siesti",
    "?",
}

def sql_str(value):
    if value is None or value == "":
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"

def sql_int(value):
    if value is None or value == "":
        return "NULL"
    try:
        return str(int(value))
    except (TypeError, ValueError):
        return "NULL"

def sql_date(value):
    if value is None or value == "":
        return "NULL"
    if hasattr(value, "strftime"):
        return "'" + value.strftime("%Y-%m-%d") + "'"
    return "NULL"

def extract_accommodation(note: str):
    if not note:
        return "NONE", note
    matched_code = "NONE"
    cleaned = note
    for pat, code in ACCOM_PATTERNS:
        if pat.search(cleaned):
            matched_code = code
            cleaned = pat.sub("", cleaned)
            break
    if matched_code != "NONE":
        for pre in PREFIX_PATTERNS:
            cleaned = pre.sub("", cleaned)
    cleaned = cleaned.strip(" ,.-")
    return matched_code, cleaned

def looks_like_name_list(note: str):
    if not note:
        return False
    low = note.lower().strip()
    if low in NON_NAME_WORDS:
        return False
    parts = re.split(r",|\s+i\s+", note)
    parts = [p.strip() for p in parts if p.strip()]
    if not parts:
        return False
    # single-token: must look like a proper noun (starts uppercase) and be short
    if len(parts) == 1:
        p = parts[0]
        return (1 < len(p) < 20
                and " " not in p
                and p[0].isupper()
                and p.lower() not in NON_NAME_WORDS)
    # multi-part: each part must be a short, name-shaped fragment
    for p in parts:
        if any(c.isdigit() for c in p) or len(p) > 30:
            return False
        if len(p.split()) > 3:
            return False
        if p.lower() in NON_NAME_WORDS:
            return False
    return True

def split_attendees(note: str):
    parts = re.split(r",|\s+i\s+", note)
    return [p.strip() for p in parts if p.strip() and p.strip().lower() not in NON_NAME_WORDS]

def main():
    wb = load_workbook(SRC, data_only=True)  # data_only resolves formulas
    ws = wb["Pozivnice"]

    statements = []
    statements.append(f"-- Generated seed from {SRC.name}")
    statements.append("BEGIN;")

    # default owner user (password: 'changeme' bcrypt hash, see README)
    owner_id = "'11111111-1111-1111-1111-111111111111'"
    # default client (tenant) that owns all seeded data; fixed UUID matches
    # db/migrations/03_multitenancy.sql so fresh-install and migrated DBs align.
    client_id = "'00000000-0000-0000-0000-0000000000c1'"
    statements.append(f"""
INSERT INTO client (id, name, slug, google_sheet_id, google_sheet_tab) VALUES
({client_id}, 'Default', 'default',
 '1gsydyLPpQH3bJoppdZoLYjlq3zKexlc-qWnuYnujeQM', 'Pozivnice')
ON CONFLICT (id) DO NOTHING;

INSERT INTO app_user (id, email, password_hash, display_name, role, is_super_admin, locale) VALUES
({owner_id},
 'owner@example.com',
 '$2b$10$wJXdVk4zM4lAGo0CaiXj8uYPOddBQcWVRnwmPMOljkIpAR9rMnK8e',  -- bcrypt('changeme')
 'Owner', 'OWNER', true, 'sr')
ON CONFLICT (email) DO NOTHING;

INSERT INTO user_client (user_id, client_id, role) VALUES
({owner_id}, {client_id}, 'OWNER')
ON CONFLICT (user_id, client_id) DO NOTHING;
""")

    inv_rows = []
    attendee_rows = []
    demoted_to_pozvan = []  # confirmed rows missing adult count — demoted for stats integrity
    unknown_statuses = []   # rows whose status didn't map to the enum

    for r in range(2, ws.max_row + 1):
        gost = ws.cell(r, 1).value
        if not gost or not str(gost).strip():
            continue
        gost = str(gost).strip()
        # skip the bottom summary row
        if gost.lower() == "planirano":
            continue

        planirano = ws.cell(r, 2).value
        status_raw = ws.cell(r, 3).value
        odrasli = ws.cell(r, 4).value
        deca = ws.cell(r, 5).value
        # F is the formula =D+E; ignore it (DB generates it)
        prognoza = ws.cell(r, 7).value
        datum = ws.cell(r, 8).value
        napomena_raw = ws.cell(r, 9).value
        napomena = str(napomena_raw).strip() if napomena_raw else ""

        status_key = norm(status_raw)
        status = STATUS_MAP.get(status_key)
        if status is None:
            unknown_statuses.append((r, status_raw))
            status = "NIJE_POZVAN"

        # chk_confirmed_requires_counts: POTVRDJEN_DOLAZAK requires adults NOT NULL.
        # If the xlsx left the headcount blank, demote to POZVAN rather than fabricate.
        if status == "POTVRDJEN_DOLAZAK" and odrasli in (None, ""):
            demoted_to_pozvan.append(gost)
            status = "POZVAN"

        # accommodation extraction
        accom, remaining_note = extract_accommodation(napomena)

        decline_reason = None
        attendees_from_note = []
        final_notes = remaining_note or None

        if status == "ODBIJENO":
            decline_reason = remaining_note or None
            final_notes = None
        elif remaining_note and looks_like_name_list(remaining_note):
            attendees_from_note = split_attendees(remaining_note)
            final_notes = None  # consumed by attendees

        inv_id = "'" + str(uuid.uuid4()) + "'"

        # Honor declined-zero-counts constraint
        if status == "ODBIJENO":
            odrasli_sql = "0"
            deca_sql = "0"
        else:
            odrasli_sql = sql_int(odrasli)
            deca_sql = sql_int(deca)

        inv_rows.append(
            f"({inv_id}, {client_id}, {sql_str(gost)}, {sql_int(planirano)}, '{status}', "
            f"{odrasli_sql}, {deca_sql}, {sql_int(prognoza)}, "
            f"{sql_date(datum)}, '{accom}', {sql_str(decline_reason)}, "
            f"{sql_str(final_notes)}, {owner_id}, {owner_id})"
        )

        # Heuristic: how many attendees are children?
        n_children = int(deca) if isinstance(deca, (int, float)) else 0
        for idx, name in enumerate(attendees_from_note):
            is_child = idx >= (len(attendees_from_note) - n_children) and n_children > 0
            attendee_rows.append(
                f"(gen_random_uuid(), {inv_id}, {sql_str(name)}, "
                f"{'TRUE' if is_child else 'FALSE'})"
            )

    statements.append("INSERT INTO invitation "
        "(id, client_id, guest_label, planned_count, status, adults, children, "
        "forecast, response_date, accommodation, decline_reason, notes, "
        "created_by, updated_by) VALUES")
    statements.append(",\n".join(inv_rows) + ";")

    if attendee_rows:
        statements.append("\nINSERT INTO attendee "
            "(id, invitation_id, full_name, is_child) VALUES")
        statements.append(",\n".join(attendee_rows) + ";")

    statements.append("\nCOMMIT;")

    OUT.write_text("\n".join(statements) + "\n", encoding="utf-8")
    print(f"Wrote {OUT}")
    print(f"  invitations: {len(inv_rows)}")
    print(f"  attendees:   {len(attendee_rows)}")
    if demoted_to_pozvan:
        print(f"  demoted POTVRDJEN_DOLAZAK → POZVAN (missing adult count): {len(demoted_to_pozvan)}")
        for name in demoted_to_pozvan:
            print(f"    - {name}")
    if unknown_statuses:
        print(
            f"WARNING: {len(unknown_statuses)} rows had unrecognised status "
            "(default 'NIJE_POZVAN' applied):",
            file=sys.stderr,
        )
        for row, raw in unknown_statuses[:20]:
            code_points = [hex(ord(c)) for c in (raw or "")]
            print(f"  row {row}: {raw!r}  codepoints={code_points}", file=sys.stderr)
        if "--strict" in sys.argv:
            print("Aborting because --strict was passed.", file=sys.stderr)
            sys.exit(1)

if __name__ == "__main__":
    main()
